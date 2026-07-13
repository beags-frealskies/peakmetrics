"""Excel history dashboard for PeakMetrics."""

from datetime import datetime
from io import BytesIO
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from openpyxl.drawing.image import (
    Image as ExcelImage,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from report_archive import (
    canonical_week_path,
    parse_date_value,
)


# PeakMetrics brand palette.
PEAK_NAVY = "#12304A"
PEAK_TEAL = "#169C9C"
PEAK_AQUA = "#63D6D0"
PEAK_FILL = "#D9F3F1"

DARK_TEXT = "#102A43"
SECONDARY_TEXT = "#486581"
LIGHT_TEXT = "#829AB1"
GRID_COLOR = "#D9E2EC"


CARD_FILL = PatternFill(
    fill_type="solid",
    fgColor="F8FAFC",
)

TABLE_HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="DDEFF0",
)

ALTERNATE_FILL = PatternFill(
    fill_type="solid",
    fgColor="F5FAFA",
)

THIN_GRAY = Side(
    style="thin",
    color="D9E2EC",
)

CARD_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=THIN_GRAY,
    bottom=THIN_GRAY,
)


def excel_color(value):
    """Remove the # for openpyxl colors."""

    return value.replace(
        "#",
        "",
    )


def draw_history_metric(
    ws,
    start_row,
    start_column,
    value,
    label,
):
    """Draw one history summary card."""

    end_column = (
        start_column + 2
    )

    ws.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=start_row + 1,
        end_column=end_column,
    )

    ws.merge_cells(
        start_row=start_row + 2,
        start_column=start_column,
        end_row=start_row + 2,
        end_column=end_column,
    )

    value_cell = ws.cell(
        row=start_row,
        column=start_column,
    )

    value_cell.value = value
    value_cell.font = Font(
        size=18,
        bold=True,
        color=excel_color(
            PEAK_NAVY
        ),
    )
    value_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    label_cell = ws.cell(
        row=start_row + 2,
        column=start_column,
    )

    label_cell.value = label
    label_cell.font = Font(
        size=9,
        color=excel_color(
            SECONDARY_TEXT
        ),
    )
    label_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for row_number in range(
        start_row,
        start_row + 3,
    ):
        for column_number in range(
            start_column,
            end_column + 1,
        ):
            cell = ws.cell(
                row=row_number,
                column=column_number,
            )

            cell.fill = CARD_FILL
            cell.border = CARD_BORDER


def format_week_chart_label(record):
    """Create a polished date label for the chart."""

    try:
        week_start = datetime.fromisoformat(
            str(record["Week Start"])
        )

        week_end = datetime.fromisoformat(
            str(record["Week End"])
        )

    except (
        KeyError,
        TypeError,
        ValueError,
    ):
        return str(
            record.get(
                "Week",
                "",
            )
        )

    if week_start.year != week_end.year:
        return (
            f"{week_start.strftime('%b')} "
            f"{week_start.day}, "
            f"{week_start.year}"
            f"–"
            f"{week_end.strftime('%b')} "
            f"{week_end.day}, "
            f"{week_end.year}"
        )

    if week_start.month != week_end.month:
        return (
            f"{week_start.strftime('%b')} "
            f"{week_start.day}"
            f"–"
            f"{week_end.strftime('%b')} "
            f"{week_end.day}"
        )

    return (
        f"{week_start.strftime('%b')} "
        f"{week_start.day}"
        f"–"
        f"{week_end.day}"
    )


def create_weekly_mileage_chart(
    weekly_history,
):
    """Create the PeakMetrics weekly-mileage chart."""

    chart_history = (
        weekly_history
        .tail(16)
        .copy()
    )

    chart_records = (
        chart_history.to_dict(
            "records"
        )
    )

    week_labels = [
        format_week_chart_label(
            record
        )
        for record in chart_records
    ]

    mileage_values = [
        float(value)
        for value
        in chart_history["Miles"]
    ]

    x_positions = list(
        range(
            len(mileage_values)
        )
    )

    figure, axis = plt.subplots(
        figsize=(12.6, 4.6),
        dpi=140,
    )

    figure.patch.set_facecolor(
        "white"
    )

    axis.set_facecolor(
        "white"
    )

    figure.subplots_adjust(
        left=0.06,
        right=0.985,
        top=0.78,
        bottom=0.22,
    )

    maximum_mileage = max(
        mileage_values
    )

    chart_ceiling = max(
        10,
        maximum_mileage * 1.22,
    )

    axis.set_ylim(
        0,
        chart_ceiling,
    )

    axis.set_axisbelow(
        True
    )

    axis.yaxis.set_major_locator(
        MaxNLocator(
            nbins=4,
            integer=True,
        )
    )

    axis.grid(
        axis="y",
        color=GRID_COLOR,
        linewidth=0.8,
    )

    axis.grid(
        axis="x",
        visible=False,
    )

    for spine in axis.spines.values():
        spine.set_visible(
            False
        )

    axis.axhline(
        0,
        color="#BCCCDC",
        linewidth=1.2,
    )

    if len(mileage_values) == 1:
        line_x = [
            -0.18,
            0.18,
        ]

        line_y = [
            mileage_values[0],
            mileage_values[0],
        ]

        axis.fill_between(
            line_x,
            line_y,
            [
                0,
                0,
            ],
            color=PEAK_FILL,
            alpha=0.7,
        )

        axis.plot(
            line_x,
            line_y,
            color=PEAK_TEAL,
            linewidth=4,
            solid_capstyle="round",
        )

        axis.scatter(
            [0],
            mileage_values,
            s=115,
            color=PEAK_AQUA,
            edgecolor="white",
            linewidth=2.5,
            zorder=5,
        )

        axis.set_xlim(
            -0.45,
            0.45,
        )

    else:
        axis.fill_between(
            x_positions,
            mileage_values,
            0,
            color=PEAK_FILL,
            alpha=0.7,
        )

        axis.plot(
            x_positions,
            mileage_values,
            color=PEAK_TEAL,
            linewidth=4,
            marker="o",
            markersize=7,
            markerfacecolor=PEAK_TEAL,
            markeredgecolor="white",
            markeredgewidth=2,
            solid_capstyle="round",
            solid_joinstyle="round",
            zorder=3,
        )

        axis.scatter(
            [x_positions[-1]],
            [mileage_values[-1]],
            s=125,
            color=PEAK_AQUA,
            edgecolor="white",
            linewidth=2.5,
            zorder=5,
        )

        axis.set_xlim(
            -0.25,
            len(x_positions) - 0.75,
        )

    axis.set_xticks(
        x_positions
    )

    axis.set_xticklabels(
        week_labels
    )

    axis.tick_params(
        axis="x",
        length=0,
        pad=12,
        labelsize=9,
        colors=SECONDARY_TEXT,
    )

    axis.tick_params(
        axis="y",
        length=0,
        pad=8,
        labelsize=9,
        colors=LIGHT_TEXT,
    )

    if len(week_labels) > 8:
        plt.setp(
            axis.get_xticklabels(),
            rotation=35,
            horizontalalignment="right",
        )

    axis.set_title(
        "Weekly Mileage",
        loc="left",
        fontsize=18,
        fontweight="bold",
        color=PEAK_NAVY,
        pad=22,
    )

    if len(chart_history) == 1:
        subtitle = (
            "Current stored training week"
        )

    else:
        subtitle = (
            f"Most recent "
            f"{len(chart_history)} weeks"
        )

    axis.text(
        0,
        1.015,
        subtitle,
        transform=axis.transAxes,
        horizontalalignment="left",
        verticalalignment="bottom",
        fontsize=10,
        color=SECONDARY_TEXT,
    )

    axis.text(
        1,
        1.015,
        (
            f"Latest: "
            f"{mileage_values[-1]:.1f} mi"
        ),
        transform=axis.transAxes,
        horizontalalignment="right",
        verticalalignment="bottom",
        fontsize=11,
        fontweight="bold",
        color=PEAK_TEAL,
    )

    if len(mileage_values) <= 8:
        label_indexes = list(
            range(
                len(mileage_values)
            )
        )

    else:
        highest_index = (
            mileage_values.index(
                maximum_mileage
            )
        )

        label_indexes = sorted(
            {
                highest_index,
                len(mileage_values) - 1,
            }
        )

    for index in label_indexes:
        axis.annotate(
            f"{mileage_values[index]:.1f}",
            (
                x_positions[index],
                mileage_values[index],
            ),
            xytext=(
                0,
                10,
            ),
            textcoords="offset points",
            horizontalalignment="center",
            verticalalignment="bottom",
            fontsize=10,
            fontweight="bold",
            color=PEAK_NAVY,
            clip_on=False,
        )

    image_buffer = BytesIO()

    figure.savefig(
        image_buffer,
        format="png",
        bbox_inches="tight",
        pad_inches=0.12,
        facecolor="white",
    )

    plt.close(
        figure
    )

    image_buffer.seek(
        0
    )

    return image_buffer


def add_workbook_hyperlink(
    cell,
    archive_path,
    sheet_name,
    display_text,
):
    """Link a cell to a sheet inside an archived workbook."""

    archive_path = Path(
        archive_path
    )

    cell.value = display_text

    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        target=(
            archive_path
            .resolve()
            .as_uri()
        ),
        location=(
            f"{sheet_name}!A1"
        ),
        display=display_text,
    )

    cell.font = Font(
        name="Arial",
        size=9,
        bold=True,
        underline="single",
        color=excel_color(
            PEAK_TEAL
        ),
    )


def create_history_sheet(
    wb,
    weekly_history,
    archive_folder,
    current_archive_path,
):
    """Create the interactive training-history sheet."""

    archive_folder = Path(
        archive_folder
    )

    current_archive_path = Path(
        current_archive_path
    )

    ws = wb.create_sheet(
        "History"
    )

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = None
    ws.sheet_properties.tabColor = (
        excel_color(
            PEAK_TEAL
        )
    )

    ws.merge_cells(
        "A1:O1"
    )

    ws["A1"] = "Training History"
    ws["A1"].font = Font(
        size=24,
        bold=True,
        color=excel_color(
            PEAK_NAVY
        ),
    )
    ws["A1"].alignment = Alignment(
        vertical="center",
    )

    ws.row_dimensions[
        1
    ].height = 34

    ws.merge_cells(
        "A2:O2"
    )

    ws["A2"] = (
        "Weekly running trends from your "
        "PeakMetrics activity database"
    )
    ws["A2"].font = Font(
        size=11,
        color=excel_color(
            SECONDARY_TEXT
        ),
    )

    if weekly_history.empty:
        ws["A4"] = (
            "No training history was found."
        )

        ws["A4"].font = Font(
            size=12,
            italic=True,
            color=excel_color(
                SECONDARY_TEXT
            ),
        )

        return

    history_summary = (
        weekly_history.attrs.get(
            "summary",
            {},
        )
    )

    total_weeks = int(
        history_summary.get(
            "Total Weeks",
            len(weekly_history),
        )
    )

    total_miles = float(
        history_summary.get(
            "Total Mileage",
            weekly_history[
                "Miles"
            ].sum(),
        )
    )

    average_week = float(
        history_summary.get(
            "Average Week",
            weekly_history[
                "Miles"
            ].mean(),
        )
    )

    current_year = int(
        history_summary.get(
            "Current Year",
            0,
        )
    )

    ytd_mileage = float(
        history_summary.get(
            "YTD Mileage",
            0,
        )
    )

    ytd_average_pace = (
        history_summary.get(
            "YTD Average Pace",
            "N/A",
        )
    )

    historical_average_pace = (
        history_summary.get(
            "Historical Average Pace",
            "N/A",
        )
    )

    best_week_index = (
        weekly_history[
            "Miles"
        ].idxmax()
    )

    best_week = (
        weekly_history.loc[
            best_week_index
        ]
    )

    draw_history_metric(
        ws,
        4,
        1,
        total_weeks,
        "Stored Weeks",
    )

    draw_history_metric(
        ws,
        4,
        5,
        f"{total_miles:.1f} mi",
        "Total Historical Mileage",
    )

    draw_history_metric(
        ws,
        4,
        9,
        f"{average_week:.1f} mi",
        "Average Weekly Mileage",
    )

    draw_history_metric(
        ws,
        4,
        13,
        (
            f"{float(best_week['Miles']):.1f} mi"
        ),
        (
            f"Highest Week: "
            f"{best_week['Week']}"
        ),
    )

    draw_history_metric(
        ws,
        8,
        1,
        f"{ytd_mileage:.1f} mi",
        f"{current_year} YTD Mileage",
    )

    draw_history_metric(
        ws,
        8,
        5,
        ytd_average_pace,
        (
            f"{current_year} "
            "YTD Average Pace"
        ),
    )

    draw_history_metric(
        ws,
        8,
        9,
        historical_average_pace,
        "Historical Average Pace",
    )

    ws.row_dimensions[
        6
    ].height = 28

    ws.row_dimensions[
        10
    ].height = 28

    mileage_chart_buffer = (
        create_weekly_mileage_chart(
            weekly_history
        )
    )

    mileage_chart = ExcelImage(
        mileage_chart_buffer
    )

    mileage_chart._peakmetrics_buffer = (
        mileage_chart_buffer
    )

    mileage_chart.width = 1120
    mileage_chart.height = 420

    ws.add_image(
        mileage_chart,
        "A14",
    )

    ws.merge_cells(
        "A37:J37"
    )

    ws["A37"] = "Weekly Details"
    ws["A37"].font = Font(
        size=14,
        bold=True,
        color=excel_color(
            PEAK_NAVY
        ),
    )

    ws.merge_cells(
        "A38:J38"
    )

    ws["A38"] = (
        "Click a week to open its Summary. "
        "Use Open Report to jump directly "
        "to the activity cards."
    )
    ws["A38"].font = Font(
        name="Arial",
        size=9,
        italic=True,
        color=excel_color(
            SECONDARY_TEXT
        ),
    )

    table_start_row = 40

    headers = [
        "Week",
        "Miles",
        "Time",
        "Average Pace",
        "Average HR",
        "Runs",
        "Workouts",
        "Long Runs",
        "Strides",
        "Open Report",
    ]

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(
            row=table_start_row,
            column=column_number,
        )

        cell.value = header
        cell.font = Font(
            name="Arial",
            bold=True,
            color=excel_color(
                PEAK_NAVY
            ),
        )
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = CARD_BORDER

    rows = weekly_history.to_dict(
        "records"
    )

    current_archive_resolved = (
        current_archive_path.resolve()
    )

    for row_number, record in enumerate(
        rows,
        start=table_start_row + 1,
    ):
        average_hr = record[
            "Average HR"
        ]

        try:
            average_hr = int(
                average_hr
            )

        except (
            TypeError,
            ValueError,
        ):
            average_hr = 0

        values = [
            record["Week"],
            float(record["Miles"]),
            record["Time"],
            record["Average Pace"],
            average_hr,
            int(record["Runs"]),
            int(record["Workouts"]),
            int(record["Long Runs"]),
            int(record["Strides"]),
            "",
        ]

        for column_number, value in enumerate(
            values,
            start=1,
        ):
            cell = ws.cell(
                row=row_number,
                column=column_number,
            )

            cell.value = value
            cell.border = CARD_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.font = Font(
                name="Arial",
                size=9,
                color=excel_color(
                    DARK_TEXT
                ),
            )

            if column_number == 2:
                cell.number_format = (
                    "0.00"
                )

            if row_number % 2 == 0:
                cell.fill = ALTERNATE_FILL

        week_start = parse_date_value(
            record["Week Start"]
        )

        week_end = parse_date_value(
            record["Week End"]
        )

        archive_path = canonical_week_path(
            archive_folder,
            week_start,
            week_end,
        )

        is_current_week = (
            archive_path.resolve()
            == current_archive_resolved
        )

        archive_available = (
            archive_path.exists()
            or is_current_week
        )

        week_cell = ws.cell(
            row=row_number,
            column=1,
        )

        report_cell = ws.cell(
            row=row_number,
            column=10,
        )

        if archive_available:
            add_workbook_hyperlink(
                week_cell,
                archive_path,
                "Summary",
                str(record["Week"]),
            )

            add_workbook_hyperlink(
                report_cell,
                archive_path,
                "Report",
                "Open Report",
            )

        else:
            week_cell.font = Font(
                name="Arial",
                size=9,
                color=excel_color(
                    DARK_TEXT
                ),
            )

            report_cell.value = (
                "Not archived"
            )
            report_cell.font = Font(
                name="Arial",
                size=9,
                italic=True,
                color=excel_color(
                    LIGHT_TEXT
                ),
            )

    final_row = (
        table_start_row
        + len(weekly_history)
    )

    ws.auto_filter.ref = (
        f"A{table_start_row}:"
        f"J{final_row}"
    )

    ws.column_dimensions[
        "A"
    ].width = 17

    ws.column_dimensions[
        "B"
    ].width = 11

    ws.column_dimensions[
        "C"
    ].width = 12

    ws.column_dimensions[
        "D"
    ].width = 15

    ws.column_dimensions[
        "E"
    ].width = 13

    ws.column_dimensions[
        "F"
    ].width = 9

    ws.column_dimensions[
        "G"
    ].width = 11

    ws.column_dimensions[
        "H"
    ].width = 11

    ws.column_dimensions[
        "I"
    ].width = 9

    ws.column_dimensions[
        "J"
    ].width = 15

    for column_number in range(
        11,
        17,
    ):
        column_letter = get_column_letter(
            column_number
        )

        ws.column_dimensions[
            column_letter
        ].width = 11