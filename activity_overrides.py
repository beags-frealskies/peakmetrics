"""Persistent manual activity-type corrections for PeakMetrics."""

import json
from pathlib import Path

from openpyxl import load_workbook


VALID_ACTIVITY_TYPES = {
    "Easy Run",
    "Long Run",
    "Workout",
    "Strides",
    "Cross Training",
    "Core",
    "Lifting",
    "Other",
}

DEFAULT_OVERRIDES_PATH = Path(
    "activity_overrides.json"
)

MAPPING_SHEET_NAME = "_ActivityData"
REPORT_SHEET_NAME = "Report"

MAX_REPORTS_TO_SCAN = 25


def load_activity_overrides(
    path=DEFAULT_OVERRIDES_PATH,
):
    """Load valid saved overrides from JSON."""

    path = Path(path)

    if not path.exists():
        return {}

    try:
        data = json.loads(
            path.read_text(
                encoding="utf-8"
            )
        )

    except (
        json.JSONDecodeError,
        OSError,
    ):
        print(
            f"Warning: Could not read {path}. "
            "Starting with no saved overrides."
        )

        return {}

    if not isinstance(data, dict):
        return {}

    overrides = {}

    for source_file, activity_type in data.items():
        source_file = str(
            source_file
        ).strip()

        activity_type = str(
            activity_type
        ).strip()

        if (
            source_file
            and activity_type
            in VALID_ACTIVITY_TYPES
        ):
            overrides[
                source_file
            ] = activity_type

    return overrides


def save_activity_overrides(
    overrides,
    path=DEFAULT_OVERRIDES_PATH,
):
    """Save valid overrides to JSON."""

    path = Path(path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    valid_overrides = {
        str(source_file): str(activity_type)
        for source_file, activity_type
        in overrides.items()
        if activity_type
        in VALID_ACTIVITY_TYPES
    }

    path.write_text(
        json.dumps(
            valid_overrides,
            indent=2,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )


def find_excel_reports(
    reports_dir,
):
    """Return recent Excel reports, newest first."""

    reports_dir = Path(
        reports_dir
    )

    preferred_reports = [
        path
        for path in reports_dir.glob(
            "PeakMetrics_*.xlsx"
        )
        if not path.name.startswith(
            "~$"
        )
    ]

    if preferred_reports:
        candidates = preferred_reports

    else:
        candidates = [
            path
            for path in reports_dir.glob(
                "*.xlsx"
            )
            if not path.name.startswith(
                "~$"
            )
        ]

    return sorted(
        candidates,
        key=lambda path: (
            path.stat().st_mtime
        ),
        reverse=True,
    )


def read_report_changes(
    report_path,
):
    """
    Read manual dropdown changes from one workbook.

    Returns None when the workbook does not contain
    the hidden PeakMetrics tracking sheet.
    """

    try:
        workbook = load_workbook(
            report_path,
            data_only=False,
        )

    except PermissionError:
        print(
            f"Skipped {report_path.name} "
            "because it is still open in Excel."
        )

        return None

    except OSError as error:
        print(
            f"Could not read "
            f"{report_path.name}: {error}"
        )

        return None

    try:
        required_sheets = {
            REPORT_SHEET_NAME,
            MAPPING_SHEET_NAME,
        }

        if not required_sheets.issubset(
            workbook.sheetnames
        ):
            return None

        report_sheet = workbook[
            REPORT_SHEET_NAME
        ]

        mapping_sheet = workbook[
            MAPPING_SHEET_NAME
        ]

        changes = []

        for row in mapping_sheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            if len(row) < 3:
                continue

            source_file = row[0]
            type_cell = row[1]
            generated_type = row[2]

            if (
                source_file is None
                or type_cell is None
                or generated_type is None
            ):
                continue

            source_file = str(
                source_file
            ).strip()

            type_cell = str(
                type_cell
            ).strip()

            generated_type = str(
                generated_type
            ).strip()

            current_type = report_sheet[
                type_cell
            ].value

            if current_type is None:
                continue

            current_type = str(
                current_type
            ).strip()

            if (
                current_type
                not in VALID_ACTIVITY_TYPES
            ):
                continue

            if current_type != generated_type:
                changes.append(
                    (
                        source_file,
                        current_type,
                    )
                )

        return changes

    finally:
        workbook.close()


def sync_overrides_from_latest_report(
    reports_dir,
    overrides_path=DEFAULT_OVERRIDES_PATH,
):
    """
    Search recent reports for saved dropdown corrections.

    Newer manual corrections take priority over older ones.

    Returns:
        overrides, number_of_changes, source_report
    """

    overrides_path = Path(
        overrides_path
    )

    overrides = load_activity_overrides(
        overrides_path
    )

    if not overrides_path.exists():
        save_activity_overrides(
            overrides,
            overrides_path,
        )

    reports = find_excel_reports(
        reports_dir
    )

    if not reports:
        return overrides, 0, None

    imported_changes = 0
    source_report = None

    # Prevent an older workbook from replacing
    # a newer correction for the same activity.
    handled_source_files = set()

    compatible_reports = 0

    for report_path in reports[
        :MAX_REPORTS_TO_SCAN
    ]:
        report_changes = read_report_changes(
            report_path
        )

        if report_changes is None:
            continue

        compatible_reports += 1

        for (
            source_file,
            activity_type,
        ) in report_changes:
            if (
                source_file
                in handled_source_files
            ):
                continue

            handled_source_files.add(
                source_file
            )

            if (
                overrides.get(source_file)
                != activity_type
            ):
                overrides[
                    source_file
                ] = activity_type

                imported_changes += 1

                if source_report is None:
                    source_report = (
                        report_path
                    )

    if compatible_reports == 0:
        print(
            "No compatible report with hidden "
            "activity tracking data was found."
        )

    if imported_changes:
        save_activity_overrides(
            overrides,
            overrides_path,
        )

    return (
        overrides,
        imported_changes,
        source_report,
    )