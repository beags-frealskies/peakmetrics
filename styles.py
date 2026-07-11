"""Shared styling for PeakMetrics Excel reports."""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# -----------------------------
# Fonts
# -----------------------------

TITLE_FONT = Font(
    size=26,
    bold=True,
)

SUBTITLE_FONT = Font(
    size=15,
)

CARD_VALUE_FONT = Font(
    size=24,
    bold=True,
)

CARD_LABEL_FONT = Font(
    size=12,
)

# -----------------------------
# Colors
# -----------------------------

CARD_FILL = PatternFill(
    fill_type="solid",
    fgColor="F8FAFC",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="EAF2FF",
)

# -----------------------------
# Borders
# -----------------------------

MEDIUM = Side(style="medium", color="D9D9D9")

CARD_BORDER = Border(
    left=MEDIUM,
    right=MEDIUM,
    top=MEDIUM,
    bottom=MEDIUM,
)

# -----------------------------
# Alignment
# -----------------------------

CENTER = Alignment(
    horizontal="center",
    vertical="center",
)

CENTER_WRAP = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)