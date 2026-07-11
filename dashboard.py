"""Dashboard drawing functions."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def draw_metric_card(ws, row, col, title, value):
    """Draw one dashboard metric."""

    fill = PatternFill(fill_type="solid", fgColor="EAF4FC")

    border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    ws.merge_cells(
        start_row=row,
        start_column=col,
        end_row=row + 3,
        end_column=col + 2,
    )

    cell = ws.cell(row=row, column=col)

    cell.value = f"{value}\n\n{title}"
    cell.font = Font(size=18, bold=True)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    cell.fill = fill

    for r in range(row, row + 4):
        for c in range(col, col + 3):
            ws.cell(r, c).fill = fill
            ws.cell(r, c).border = border