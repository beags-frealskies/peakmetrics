"""PeakMetrics Excel activity cards."""

from copy import copy
from datetime import datetime

import pandas as pd
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.datavalidation import DataValidation

from config import CONFIG


ACTIVITY_TYPES = [
    "Easy Run",
    "Long Run",
    "Workout",
    "Strides",
    "Cross Training",
    "Core",
    "Lifting",
    "Other",
]


# PeakMetrics brand colors.
PEAK_NAVY = "12304A"
PEAK_TEAL = "169C9C"
PEAK_AQUA = "63D6D0"

DARK_TEXT = "102A43"
SECONDARY_TEXT = "486581"
LIGHT_TEXT = "829AB1"

WHITE = "FFFFFF"
CARD_BACKGROUND = "FFFFFF"
METRIC_BACKGROUND = "F5FAFA"
TABLE_HEADER = "DDEFF0"
ALTERNATE_ROW = "F7FAFC"

BORDER_COLOR = "D9E2EC"


RUN_TYPE_STYLES = {
    "Easy Run": {
        "accent": "169C9C",
        "soft": "E8F7F5",
    },
    "Long Run": {
        "accent": "12304A",
        "soft": "EAF0F5",
    },
    "Workout": {
        "accent": "D58B17",
        "soft": "FCF4E5",
    },
    "Strides": {
        "accent": "7257A8",
        "soft": "F1ECF8",
    },
    "Cross Training": {
        "accent": "2878B5",
        "soft": "EAF3FA",
    },
    "Core": {
        "accent": "4B7F52",
        "soft": "EDF5EE",
    },
    "Lifting": {
        "accent": "5B667A",
        "soft": "EEF0F3",
    },
    "Other": {
        "accent": "6B7280",
        "soft": "F2F4F7",
    },
}


THIN_SIDE = Side(
    style="thin",
    color=BORDER_COLOR,
)

MEDIUM_SIDE = Side(
    style="medium",
    color=BORDER_COLOR,
)

NO_SIDE = Side(
    style=None,
)

METRIC_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


def is_missing(value):
    """Return True for blank, None, or NaN values."""

    if value is None:
        return True

    if isinstance(
        value,
        (
            list,
            tuple,
            dict,
        ),
    ):
        return False

    try:
        return bool(
            pd.isna(value)
        )

    except (
        TypeError,
        ValueError,
    ):
        return False


def numeric_text(
    value,
    suffix="",
    decimals=0,
):
    """Format a numeric metric with its unit."""

    if is_missing(value):
        return "—"

    try:
        number = float(value)

    except (
        TypeError,
        ValueError,
    ):
        return str(value)

    if decimals == 0:
        body = f"{number:,.0f}"
    else:
        body = f"{number:,.{decimals}f}"

    if suffix:
        return f"{body} {suffix}"

    return body


def text_value(value):
    """Return readable text for an optional value."""

    if is_missing(value):
        return "—"

    value = str(value).strip()

    if not value:
        return "—"

    return value


def activity_heading(run):
    """Create the activity date and start-time heading."""

    start_time = run.get(
        "Start Time"
    )

    if (
        start_time is not None
        and hasattr(
            start_time,
            "strftime",
        )
    ):
        day_text = start_time.strftime(
            "%A, %B"
        )

        time_text = (
            start_time.strftime(
                "%I:%M %p"
            )
            .lstrip("0")
        )

        return (
            f"{day_text} "
            f"{start_time.day}  •  "
            f"{time_text}"
        )

    date_value = run.get(
        "Date",
        "",
    )

    time_value = run.get(
        "Time of Day",
        "",
    )

    try:
        parsed_date = datetime.fromisoformat(
            str(date_value)
        )

        date_text = (
            f"{parsed_date.strftime('%A, %B')} "
            f"{parsed_date.day}"
        )

    except ValueError:
        date_text = str(
            date_value
        )

    if time_value:
        return (
            f"{date_text}  •  "
            f"{time_value}"
        )

    return date_text


def format_pace(value):
    """Format pace without duplicating the unit."""

    value = text_value(
        value
    )

    if value == "—":
        return value

    return value.replace(
        "/mi",
        "",
    )


def style_range(
    ws,
    start_row,
    end_row,
    start_column,
    end_column,
    fill=None,
    border=None,
):
    """Apply shared styling across a cell range."""

    for row_number in range(
        start_row,
        end_row + 1,
    ):
        for column_number in range(
            start_column,
            end_column + 1,
        ):
            cell = ws.cell(
                row=row_number,
                column=column_number,
            )

            if fill is not None:
                cell.fill = fill

            if border is not None:
                cell.border = border


def prepare_card_background(
    ws,
    start_row,
    end_row,
):
    """Create the clean outlined card background."""

    background_fill = PatternFill(
        fill_type="solid",
        fgColor=CARD_BACKGROUND,
    )

    for row_number in range(
        start_row,
        end_row + 1,
    ):
        for column_number in range(
            1,
            17,
        ):
            cell = ws.cell(
                row=row_number,
                column=column_number,
            )

            cell.fill = background_fill

            cell.border = Border(
                left=(
                    THIN_SIDE
                    if column_number == 1
                    else NO_SIDE
                ),
                right=(
                    THIN_SIDE
                    if column_number == 16
                    else NO_SIDE
                ),
                top=(
                    THIN_SIDE
                    if row_number == start_row
                    else NO_SIDE
                ),
                bottom=(
                    MEDIUM_SIDE
                    if row_number == end_row
                    else NO_SIDE
                ),
            )


def draw_metric_tile(
    ws,
    value_row,
    label_row,
    start_column,
    value,
    label,
):
    """Draw one compact activity metric."""

    end_column = start_column + 1

    ws.merge_cells(
        start_row=value_row,
        start_column=start_column,
        end_row=value_row,
        end_column=end_column,
    )

    ws.merge_cells(
        start_row=label_row,
        start_column=start_column,
        end_row=label_row,
        end_column=end_column,
    )

    style_range(
        ws,
        value_row,
        label_row,
        start_column,
        end_column,
        fill=PatternFill(
            fill_type="solid",
            fgColor=METRIC_BACKGROUND,
        ),
        border=METRIC_BORDER,
    )

    value_cell = ws.cell(
        row=value_row,
        column=start_column,
    )

    value_cell.value = value
    value_cell.font = Font(
        name="Arial",
        size=11,
        bold=True,
        color=DARK_TEXT,
    )
    value_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    label_cell = ws.cell(
        row=label_row,
        column=start_column,
    )

    label_cell.value = label
    label_cell.font = Font(
        name="Arial",
        size=8,
        color=SECONDARY_TEXT,
    )
    label_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )


def merge_lap_cells(
    ws,
    row_number,
    start_column,
    end_column,
    value,
    fill,
    font,
):
    """Create one merged lap-table cell."""

    if start_column != end_column:
        ws.merge_cells(
            start_row=row_number,
            start_column=start_column,
            end_row=row_number,
            end_column=end_column,
        )

    style_range(
        ws,
        row_number,
        row_number,
        start_column,
        end_column,
        fill=fill,
        border=METRIC_BORDER,
    )

    cell = ws.cell(
        row=row_number,
        column=start_column,
    )

    cell.value = value
    cell.font = font
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )


def draw_lap_table(
    ws,
    run,
    start_row,
):
    """Draw the activity lap-splits table."""

    laps = run.get(
        "Laps"
    ) or []

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=TABLE_HEADER,
    )

    ws.merge_cells(
        start_row=start_row + 2,
        start_column=9,
        end_row=start_row + 2,
        end_column=16,
    )

    style_range(
        ws,
        start_row + 2,
        start_row + 2,
        9,
        16,
        fill=header_fill,
        border=METRIC_BORDER,
    )

    lap_title = ws.cell(
        row=start_row + 2,
        column=9,
    )

    lap_title.value = "Lap Splits"
    lap_title.font = Font(
        name="Arial",
        size=10,
        bold=True,
        color=PEAK_NAVY,
    )
    lap_title.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    column_font = Font(
        name="Arial",
        size=9,
        bold=True,
        color=SECONDARY_TEXT,
    )

    merge_lap_cells(
        ws,
        start_row + 3,
        9,
        9,
        "Lap",
        header_fill,
        column_font,
    )

    merge_lap_cells(
        ws,
        start_row + 3,
        10,
        11,
        "Distance",
        header_fill,
        column_font,
    )

    merge_lap_cells(
        ws,
        start_row + 3,
        12,
        13,
        "Pace",
        header_fill,
        column_font,
    )

    merge_lap_cells(
        ws,
        start_row + 3,
        14,
        15,
        "Time",
        header_fill,
        column_font,
    )

    merge_lap_cells(
        ws,
        start_row + 3,
        16,
        16,
        "Avg HR",
        header_fill,
        column_font,
    )

    if not laps:
        ws.merge_cells(
            start_row=start_row + 4,
            start_column=9,
            end_row=start_row + 5,
            end_column=16,
        )

        style_range(
            ws,
            start_row + 4,
            start_row + 5,
            9,
            16,
            fill=PatternFill(
                fill_type="solid",
                fgColor=WHITE,
            ),
            border=METRIC_BORDER,
        )

        empty_cell = ws.cell(
            row=start_row + 4,
            column=9,
        )

        empty_cell.value = (
            "No lap data available"
        )
        empty_cell.font = Font(
            name="Arial",
            size=9,
            italic=True,
            color=LIGHT_TEXT,
        )
        empty_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        return

    data_font = Font(
        name="Arial",
        size=9,
        color=DARK_TEXT,
    )

    bold_data_font = Font(
        name="Arial",
        size=9,
        bold=True,
        color=DARK_TEXT,
    )

    for lap_index, lap in enumerate(
        laps,
        start=1,
    ):
        row_number = (
            start_row
            + 3
            + lap_index
        )

        row_fill = PatternFill(
            fill_type="solid",
            fgColor=(
                ALTERNATE_ROW
                if lap_index % 2 == 0
                else WHITE
            ),
        )

        lap_number = lap.get(
            "Lap",
            lap_index,
        )

        lap_miles = numeric_text(
            lap.get(
                "Miles"
            ),
            "mi",
            decimals=2,
        )

        lap_pace = format_pace(
            lap.get(
                "Pace"
            )
        )

        lap_time = text_value(
            lap.get(
                "Time"
            )
        )

        lap_hr = numeric_text(
            lap.get(
                "Avg HR"
            ),
            "bpm",
        )

        merge_lap_cells(
            ws,
            row_number,
            9,
            9,
            lap_number,
            row_fill,
            bold_data_font,
        )

        merge_lap_cells(
            ws,
            row_number,
            10,
            11,
            lap_miles,
            row_fill,
            data_font,
        )

        merge_lap_cells(
            ws,
            row_number,
            12,
            13,
            lap_pace,
            row_fill,
            data_font,
        )

        merge_lap_cells(
            ws,
            row_number,
            14,
            15,
            lap_time,
            row_fill,
            data_font,
        )

        merge_lap_cells(
            ws,
            row_number,
            16,
            16,
            lap_hr,
            row_fill,
            data_font,
        )

        ws.row_dimensions[
            row_number
        ].height = 21


def draw_activity_card(
    ws,
    run,
    start_row,
    type_validation,
):
    """Draw one complete PeakMetrics activity card."""

    laps = run.get(
        "Laps"
    ) or []

    end_row = max(
        start_row + 6,
        start_row + 3 + len(laps),
    )

    prepare_card_background(
        ws,
        start_row,
        end_row,
    )

    run_type = text_value(
        run.get(
            "Run Type",
            "Other",
        )
    )

    if run_type not in ACTIVITY_TYPES:
        run_type = "Other"

    run_style = RUN_TYPE_STYLES.get(
        run_type,
        RUN_TYPE_STYLES["Other"],
    )

    accent_fill = PatternFill(
        fill_type="solid",
        fgColor=run_style["accent"],
    )

    soft_fill = PatternFill(
        fill_type="solid",
        fgColor=run_style["soft"],
    )

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=8,
    )

    ws.merge_cells(
        start_row=start_row,
        start_column=9,
        end_row=start_row,
        end_column=16,
    )

    style_range(
        ws,
        start_row,
        start_row,
        1,
        16,
        fill=accent_fill,
        border=Border(
            top=Side(
                style="medium",
                color=run_style["accent"],
            ),
            bottom=Side(
                style="medium",
                color=run_style["accent"],
            ),
        ),
    )

    heading_cell = ws.cell(
        row=start_row,
        column=1,
    )

    heading_cell.value = activity_heading(
        run
    )
    heading_cell.font = Font(
        name="Arial",
        size=11,
        bold=True,
        color=WHITE,
    )
    heading_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    type_cell = ws.cell(
        row=start_row,
        column=9,
    )

    type_cell.value = run_type
    type_cell.font = Font(
        name="Arial",
        size=10,
        bold=True,
        color=WHITE,
    )
    type_cell.alignment = Alignment(
        horizontal="right",
        vertical="center",
    )

    type_validation.add(
        type_cell
    )

    ws.row_dimensions[
        start_row
    ].height = 27

    style_range(
        ws,
        start_row + 1,
        start_row + 1,
        1,
        16,
        fill=soft_fill,
        border=Border(
            bottom=THIN_SIDE,
        ),
    )

    ws.row_dimensions[
        start_row + 1
    ].height = 5

    metrics = [
        (
            numeric_text(
                run.get("Miles"),
                "mi",
                decimals=2,
            ),
            "Distance",
        ),
        (
            text_value(
                run.get("Time")
            ),
            "Time",
        ),
        (
            format_pace(
                run.get(
                    "Pace (/mi)"
                )
            ),
            "Average Pace",
        ),
        (
            numeric_text(
                run.get("Avg HR"),
                "bpm",
            ),
            "Average HR",
        ),
        (
            numeric_text(
                run.get(
                    "Elevation (ft)"
                ),
                "ft",
            ),
            "Elevation Gain",
        ),
        (
            numeric_text(
                run.get("Power (W)"),
                "W",
            ),
            "Average Power",
        ),
        (
            numeric_text(
                run.get("Max HR"),
                "bpm",
            ),
            "Maximum HR",
        ),
        (
            numeric_text(
                run.get(
                    "Cadence (spm)"
                ),
                "spm",
            ),
            "Cadence",
        ),
    ]

    metric_columns = [
        1,
        3,
        5,
        7,
    ]

    for index, (
        value,
        label,
    ) in enumerate(
        metrics[:4]
    ):
        draw_metric_tile(
            ws,
            start_row + 2,
            start_row + 3,
            metric_columns[index],
            value,
            label,
        )

    for index, (
        value,
        label,
    ) in enumerate(
        metrics[4:]
    ):
        draw_metric_tile(
            ws,
            start_row + 4,
            start_row + 5,
            metric_columns[index],
            value,
            label,
        )

    ws.row_dimensions[
        start_row + 2
    ].height = 24

    ws.row_dimensions[
        start_row + 3
    ].height = 18

    ws.row_dimensions[
        start_row + 4
    ].height = 24

    ws.row_dimensions[
        start_row + 5
    ].height = 18

    draw_lap_table(
        ws,
        run,
        start_row,
    )

    for row_number in range(
        start_row + 6,
        end_row + 1,
    ):
        if (
            ws.row_dimensions[
                row_number
            ].height
            is None
        ):
            ws.row_dimensions[
                row_number
            ].height = 21

    return end_row


def apply_report_font(
    ws,
    final_row,
):
    """Set every visible Report-sheet font to Arial."""

    for row in ws.iter_rows(
        min_row=1,
        max_row=final_row,
        min_col=1,
        max_col=16,
    ):
        for cell in row:
            if cell.value is None:
                continue

            updated_font = copy(
                cell.font
            )

            updated_font.name = "Arial"
            cell.font = updated_font


def create_report_sheet(
    wb,
    df,
):
    """Create the branded chronological activity report."""

    ws = wb.create_sheet(
        "Report"
    )

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 80
    ws.freeze_panes = None
    ws.sheet_properties.tabColor = PEAK_NAVY

    ws.page_setup.orientation = (
        "landscape"
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    ws.merge_cells(
        "A1:P1"
    )

    ws["A1"] = "Weekly Activities"
    ws["A1"].font = Font(
        name="Arial",
        size=24,
        bold=True,
        color=PEAK_NAVY,
    )
    ws["A1"].alignment = Alignment(
        vertical="center",
    )

    ws.row_dimensions[1].height = 34

    ws.merge_cells(
        "A2:P2"
    )

    ws["A2"] = (
        f"{CONFIG.athlete_name}  •  "
        f"{CONFIG.team}  •  "
        f"{len(df)} activities"
    )
    ws["A2"].font = Font(
        name="Arial",
        size=10,
        color=SECONDARY_TEXT,
    )
    ws["A2"].alignment = Alignment(
        vertical="center",
    )

    ws.row_dimensions[2].height = 21

    ws.merge_cells(
        "A3:P3"
    )

    style_range(
        ws,
        3,
        3,
        1,
        16,
        fill=PatternFill(
            fill_type="solid",
            fgColor=PEAK_TEAL,
        ),
    )

    ws.row_dimensions[3].height = 4

    type_validation = DataValidation(
        type="list",
        formula1=(
            '"'
            + ",".join(
                ACTIVITY_TYPES
            )
            + '"'
        ),
        allow_blank=False,
    )

    type_validation.showInputMessage = False
    type_validation.showErrorMessage = True
    type_validation.errorTitle = (
        "Invalid activity type"
    )
    type_validation.error = (
        "Choose an activity type from "
        "the dropdown list."
    )

    ws.add_data_validation(
        type_validation
    )

    start_row = 4

    for run in df.to_dict(
        "records"
    ):
        end_row = draw_activity_card(
            ws,
            run,
            start_row,
            type_validation,
        )

        start_row = end_row + 3

    column_widths = {
        "A": 10,
        "B": 10,
        "C": 10,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 7,
        "J": 8,
        "K": 8,
        "L": 8,
        "M": 8,
        "N": 8,
        "O": 8,
        "P": 11,
    }

    for (
        column_letter,
        width,
    ) in column_widths.items():
        ws.column_dimensions[
            column_letter
        ].width = width

    final_row = max(
        3,
        start_row - 3,
    )

    apply_report_font(
        ws,
        final_row,
    )

    ws.print_title_rows = "1:3"
    ws.print_area = (
        f"A1:P{final_row}"
    )

    ws.sheet_view.selection[
        0
    ].activeCell = "A1"

    ws.sheet_view.selection[
        0
    ].sqref = "A1"