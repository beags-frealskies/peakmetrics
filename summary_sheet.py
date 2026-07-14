"""PeakMetrics Excel summary dashboard."""

from datetime import datetime
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from config import CONFIG
from mileage_chart import create_daily_mileage_chart


# PeakMetrics brand palette.
PEAK_NAVY = "#12304A"
PEAK_TEAL = "#169C9C"
PEAK_AQUA = "#63D6D0"
PEAK_FILL = "#D9F3F1"

DARK_TEXT = "#102A43"
SECONDARY_TEXT = "#486581"
LIGHT_TEXT = "#829AB1"
GRID_COLOR = "#D9E2EC"

CARD_FILL = PatternFill(
    fill_type="solid",
    fgColor="F8FAFC",
)

INFO_FILL = PatternFill(
    fill_type="solid",
    fgColor="F5FAFA",
)

ACCENT_FILL = PatternFill(
    fill_type="solid",
    fgColor="169C9C",
)

THIN_GRAY = Side(
    style="thin",
    color="D9E2EC",
)

TEAL_TOP = Side(
    style="medium",
    color="169C9C",
)

CARD_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=TEAL_TOP,
    bottom=THIN_GRAY,
)

INFO_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=THIN_GRAY,
    bottom=THIN_GRAY,
)


def format_week_range(
    start_value,
    end_value,
):
    """Create a clean display date range."""

    try:
        start_date = datetime.fromisoformat(
            str(start_value)
        )

        end_date = datetime.fromisoformat(
            str(end_value)
        )

    except ValueError:
        return (
            f"{start_value} – {end_value}"
        )

    if start_date.year != end_date.year:
        return (
            f"{start_date.strftime('%b')} "
            f"{start_date.day}, "
            f"{start_date.year} – "
            f"{end_date.strftime('%b')} "
            f"{end_date.day}, "
            f"{end_date.year}"
        )

    if start_date.month != end_date.month:
        return (
            f"{start_date.strftime('%b')} "
            f"{start_date.day} – "
            f"{end_date.strftime('%b')} "
            f"{end_date.day}, "
            f"{end_date.year}"
        )

    return (
        f"{start_date.strftime('%b')} "
        f"{start_date.day}–"
        f"{end_date.day}, "
        f"{end_date.year}"
    )


def pace_display(value):
    """Ensure pace values show their unit."""

    value = str(value)

    if value == "N/A":
        return value

    if "/mi" in value:
        return value

    return f"{value}/mi"


def draw_info_card(
    ws,
    start_column,
    end_column,
    label,
    value,
):
    """Draw one athlete-information card."""

    ws.merge_cells(
        start_row=4,
        start_column=start_column,
        end_row=4,
        end_column=end_column,
    )

    ws.merge_cells(
        start_row=5,
        start_column=start_column,
        end_row=6,
        end_column=end_column,
    )

    label_cell = ws.cell(
        row=4,
        column=start_column,
    )

    label_cell.value = label.upper()
    label_cell.font = Font(
        size=8,
        bold=True,
        color=SECONDARY_TEXT.replace(
            "#",
            "",
        ),
    )
    label_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    value_cell = ws.cell(
        row=5,
        column=start_column,
    )

    value_cell.value = value
    value_cell.font = Font(
        size=12,
        bold=True,
        color=PEAK_NAVY.replace(
            "#",
            "",
        ),
    )
    value_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    for row_number in range(
        4,
        7,
    ):
        for column_number in range(
            start_column,
            end_column + 1,
        ):
            cell = ws.cell(
                row=row_number,
                column=column_number,
            )

            cell.fill = INFO_FILL
            cell.border = INFO_BORDER


def draw_metric_card(
    ws,
    start_row,
    start_column,
    value,
    label,
):
    """Draw one dashboard metric card."""

    end_column = start_column + 3
    end_row = start_row + 3

    ws.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=start_row + 1,
        end_column=end_column,
    )

    ws.merge_cells(
        start_row=start_row + 2,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )

    value_cell = ws.cell(
        row=start_row,
        column=start_column,
    )

    value_cell.value = value
    value_cell.font = Font(
        size=18,
        bold=True,
        color=PEAK_NAVY.replace(
            "#",
            "",
        ),
    )
    value_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    label_cell = ws.cell(
        row=start_row + 2,
        column=start_column,
    )

    label_cell.value = label
    label_cell.font = Font(
        size=9,
        color=SECONDARY_TEXT.replace(
            "#",
            "",
        ),
    )
    label_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for row_number in range(
        start_row,
        end_row + 1,
    ):
        for column_number in range(
            start_column,
            end_column + 1,
        ):
            cell = ws.cell(
                row=row_number,
                column=column_number,
            )

            cell.fill = CARD_FILL
            cell.border = CARD_BORDER


def create_summary_sheet(
    wb,
    df,
    summary,
    daily_mileage,
):
    """Create the complete Summary dashboard."""

    ws = wb.active
    ws.title = "Summary"

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = None
    ws.sheet_properties.tabColor = "169C9C"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.merge_cells(
        "A1:P1"
    )

    ws["A1"] = "PeakMetrics"
    ws["A1"].font = Font(
        size=26,
        bold=True,
        color="12304A",
    )
    ws["A1"].alignment = Alignment(
        vertical="center",
    )

    ws.row_dimensions[1].height = 35

    ws.merge_cells(
        "A2:P2"
    )

    ws["A2"] = "Weekly Training Overview"
    ws["A2"].font = Font(
        size=12,
        color="486581",
    )

    ws["A2"].alignment = Alignment(
        vertical="center",
    )

    ws.merge_cells(
        "A3:P3"
    )

    ws["A3"].fill = ACCENT_FILL
    ws.row_dimensions[3].height = 4

    dates = sorted(
        str(value)
        for value in df["Date"]
    )

    week_display = format_week_range(
        dates[0],
        dates[-1],
    )

    draw_info_card(
        ws,
        1,
        5,
        "Athlete",
        CONFIG.athlete_name,
    )

    draw_info_card(
        ws,
        6,
        10,
        "Team",
        CONFIG.team,
    )

    draw_info_card(
        ws,
        11,
        16,
        "Reporting Week",
        week_display,
    )

    # Primary weekly metrics.
    draw_metric_card(
        ws,
        8,
        1,
        f"{summary.mileage:.1f} mi",
        "Weekly Mileage",
    )

    draw_metric_card(
        ws,
        8,
        5,
        summary.weekly_time,
        "Weekly Time",
    )

    draw_metric_card(
        ws,
        8,
        9,
        pace_display(
            summary.average_pace
        ),
        "Average Pace",
    )

    draw_metric_card(
        ws,
        8,
        13,
        summary.runs,
        "Runs",
    )

    # Supporting weekly metrics.
    draw_metric_card(
        ws,
        13,
        1,
        f"{summary.average_hr} bpm",
        "Average Heart Rate",
    )

    draw_metric_card(
        ws,
        13,
        5,
        f"{summary.max_hr} bpm",
        "Maximum Heart Rate",
    )

    draw_metric_card(
        ws,
        13,
        9,
        f"{summary.average_power} W",
        "Average Power",
    )

    draw_metric_card(
        ws,
        13,
        13,
        f"{summary.elevation_gain:,} ft",
        "Elevation Gain",
    )

    if daily_mileage.empty:
        ws.merge_cells(
            "A20:P23"
        )

        ws["A20"] = (
            "No running mileage was found "
            "for this reporting period."
        )

        ws["A20"].font = Font(
            size=12,
            italic=True,
            color="486581",
        )

        ws["A20"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    else:
        chart_buffer = (
            create_daily_mileage_chart(
                daily_mileage,
                summary.mileage,
            )
        )

        chart_image = ExcelImage(
            chart_buffer
        )

        # Keep the image buffer alive until
        # the workbook finishes saving.
        chart_image._peakmetrics_buffer = (
            chart_buffer
        )

        chart_image.width = 1120
        chart_image.height = 385

        ws.add_image(
            chart_image,
            "A19",
        )

    for row_number in range(
        19,
        40,
    ):
        ws.row_dimensions[
            row_number
        ].height = 18

    for column_number in range(
        1,
        17,
    ):
        column_letter = get_column_letter(
            column_number
        )

        ws.column_dimensions[
            column_letter
        ].width = 9.5

    ws.sheet_view.selection[0].activeCell = "A1"
    ws.sheet_view.selection[0].sqref = "A1"